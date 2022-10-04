"""
https://openpyxl.readthedocs.io/en/stable/
pip install openpyxl
pipenv install openpyxl
"""
import openpyxl

pedidos = openpyxl.load_workbook('pedidos.xlsx')
nome_planilhas = pedidos.sheetnames
planilha1 = pedidos['Página1']

print('\n--------------------------------------------------------------------------------------------------------\n')

print(planilha1['a1'].value)
print(planilha1['a2'].value)
print(planilha1['a3'].value)
print(planilha1['a4'].value)

print('\n--------------------------------------------------------------------------------------------------------\n')

print(planilha1['b1'].value)
print(planilha1['b2'].value)
print(planilha1['b3'].value)
print(planilha1['b4'].value)

print('\n--------------------------------------------------------------------------------------------------------\n')

print(planilha1['c1'].value)
print(planilha1['c2'].value)
print(planilha1['c3'].value)
print(planilha1['c4'].value)

print('\n--------------------------------------------------------------------------------------------------------\n')

for campo in planilha1['b']:
    print(campo.value)

print('\n--------------------------------------------------------------------------------------------------------\n')

for linha in planilha1['a1:c2']:
    for coluna in linha:
        print(coluna.value)

print('\n--------------------------------------------------------------------------------------------------------\n')

for linha in planilha1:
    if linha[0].value is not None:
        print(linha[0].value, end=' ')
    if linha[1].value is not None:
        print(linha[1].value, end=' ')
    if linha[2].value is not None:
        print(linha[2].value)

print('\n--------------------------------------------------------------------------------------------------------\n')

print('Criação da planilha: nova_planilha1')

from random import uniform
planilha2 = pedidos['Página1']

for linha in range(5, 16):
    numero_pedido = linha - 1
    planilha2.cell(linha, 1).value = numero_pedido
    planilha2.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)
    planilha2.cell(linha, 3).value = preco


pedidos.save('nova_planilha1.xlsx')

print('\n--------------------------------------------------------------------------------------------------------\n')

print('Criação da planilha: nova_planilha2')

planilha3 = openpyxl.Workbook()
planilha3.create_sheet('Planilha1', 0)
planilha3.create_sheet('Planilha2', 1)

planilha3_1 = planilha3['Planilha1']
planilha3_2 = planilha3['Planilha2']

for linha in range(1, 11):
    numero_pedido = linha - 1
    planilha3_1.cell(linha, 1).value = numero_pedido
    planilha3_1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)
    planilha3_1.cell(linha, 3).value = preco

for linha in range(1, 11):
    planilha3_2.cell(linha, 1).value = f'Italo {linha} {round(uniform(10, 100), 2)}'
    planilha3_2.cell(linha, 2).value = f'Silva {linha} {round(uniform(10, 100), 2)}'
    planilha3_2.cell(linha, 3).value = f'Fernandes {linha} {round(uniform(10, 100), 2)}'

planilha3.save('nova_planilha2.xlsx')

print('\n--------------------------------------------------------------------------------------------------------\n')
